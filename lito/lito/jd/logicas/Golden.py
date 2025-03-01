'''Generacion de reportes '''


from openpyxl import Workbook
from flask import request,  make_response
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as botcito
from typing import Generic, TypeVar
from datetime import datetime

from datos.cortas import meses
from ayudante import Celebridades, Interacciones
from datos import Consultas, Pleyades, RetencionesQ, Reportes
from logicas import Providencias, Informe, Cadena, Fotones


class InformeRevreporte(Pleyades.Abayo):
    ''' Reporte de Revision: '''

    xjd = ''
    nav = ''

    def __init__(self, db):
        super().__init__(db.uf)
        self.db = db
        self.uf = self.db.uf
        self.reporte_pie = False
        self.globales = 0
        self.df = botcito.DataFrame()
        self.procedencia = 'externa'
        if isinstance(self.xjd, str):
            self.procedencia = 'interna'
        match self.procedencia:
            case "externa":
                self.jd = self.xjd

            case "interna":
                self.jd = self.db.uf.pi
                self.nav = self.db.uf.navegante

        self._sql = RetencionesQ.Terceros(self.jd)

    def get_hoja_cruces(self, worksheet):
        '''para interno y externo'''
        info = Informe.Revision(self.db)
        self.df = info.get_archivo_ir(self._sql)

        self.df.columns = [c.replace("\t", "_") for c in self.df.columns]
        rows = dataframe_to_rows(self.df)
        worksheet.sheet_properties.tabColor = '1072BA'
        # worksheet.freeze_panes = 'L2'
        worksheet.merge_cells('E4:I4')
        worksheet.merge_cells('E5:I5')
        worksheet.merge_cells('F6:I6')
        worksheet.merge_cells('F7:I7')
        worksheet.merge_cells('F8:I8')
        if self.globales == 0:
            self.nav.nombre_analista = self.nav.nombre
            self.nav.nombre_supervisor = ''

        self.escribir_encabezado(worksheet,  '', self._sql.jd)
        ubicar_en_fila, ubicar_en_col = 10, 4
        for r_idx, row in enumerate(rows, 1):
            for c_idx, valor in enumerate(row, 1):
                if c_idx < 2:
                    continue
                if r_idx == 1:
                    self.escribir_celda(worksheet, r_idx+ubicar_en_fila, c_idx - 1 + ubicar_en_col, valor, centrado=True, negrita=True, fondo=self.uf.sri, colorido=self.uf.blanco)
                else:
                    self.escribir_celda(worksheet, r_idx - 1 + ubicar_en_fila, c_idx - 1 + ubicar_en_col, valor, centrado=False, negrita=False, fondo=self.uf.peach, colorido=self.uf.negro)

        self.auto_ajustar(worksheet)
        worksheet.sheet_view.showGridLines = False
        worksheet.sheet_view.zoomScale = 80
        if self.reporte_pie:
            self.escribir_footer(worksheet, len(self.df.index), 6, self.nav)
        return worksheet

    def get_informe(self):
        '''informe generico dependiendo de la clase'''
        self._sql.jd.procedencia = 'interna'
        workbook = Workbook()
        workbook.remove(workbook.active)
        worksheet = workbook.create_sheet(title='Cruce Retenciones', index=1)
        self.reporte_pie = True
        worksheet = self.get_hoja_cruces(worksheet)
        worksheet.title = f'Cruce Retenciones ({len(self.df.index)})'
        return workbook


class Cadenareporte(Pleyades.Abayo):
    ''' Reporte de Cadenas: '''
    xjd = ''
    nav = ''

    def __init__(self, db):
        '''constructor principal'''
        super().__init__(db.uf)
        self.db = db
        self.uf = self.db.uf
        self.reporte_pie = True
        self.globales = 0
        self.df = botcito.DataFrame()
        self.procedencia = 'externa'
        if isinstance(self.xjd, str):
            self.procedencia = 'interna'
        match self.procedencia:
            case "externa":
                self.jd = self.xjd

            case "interna":
                self.jd = self.db.uf.pi
                self.nav = self.db.uf.navegante

        self._sql = Reportes.Globales(self.jd)

    def get_hoja_cadena(self, worksheet):
        '''hoja principal de la cadena '''
        worksheet.sheet_properties.tabColor = self.uf.azul
        # worksheet.freeze_panes = 'L2'
        worksheet.merge_cells('E4:I4')
        worksheet.merge_cells('E5:I5')
        worksheet.merge_cells('F6:I6')
        worksheet.merge_cells('F7:I7')
        worksheet.merge_cells('F8:I8')
        cade = Cadena.Iva(self.db)
        df_parcial = cade.get_cadena_procesada(self._sql)
        self.df = df_parcial.copy()
        df_parcial = df_parcial.drop(['camino'], axis=1).copy()
        df_parcial = df_parcial.fillna(0)
        df_parcial_tr = df_parcial.T
        df_parcial_tr.columns = \
            [f"PERIODO {i+1}" for i in range(df_parcial_tr.shape[1])]
        self.escribir_encabezado(worksheet,  '', self._sql.jd)
        ubicar_en_fila, ubicar_en_col = 10, 4
        c_idx = 0
        columnas = ['DESCRIPCION/PERIODO', '', 'AÑO FISCAL', 'MES FISCAL',
                    'IMPUESTO CAUSADO', 'CREDITO TRIBUTARIO MES ACTUAL',
                    'SALDO CREDITO TRIBUTARIO POR ADQUISICIONES MES ANTERIOR',
                    'SALDO CREDITO TRIBUTARIO POR RETENCIONES MES ANTERIOR',
                    'RETENCIONES EN LA FUENTE DE IVA QUE LE HAN SIDO EFECTUADAS',
                    'SALDO CREDITO TRIBUTARIO X ADQUISICIONES PROXIMO MES',
                    'SALDO CREDITO TRIBUTARIO X RETENCIONES PROXIMO MES',
                    'TOTAL IMPUESTO A PAGAR', 'RETENCIONES DE IVA A DEVOLVER']

        rows = dataframe_to_rows(df_parcial_tr)
        num_columnas = len(columnas)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, valor in enumerate(row, 1):
                texto = valor 
                negrita = False
                if c_idx == 1 and r_idx <= num_columnas:
                    texto = columnas[r_idx-1]
                    negrita = True

                self.escribir_celda(worksheet, r_idx - 1 + ubicar_en_fila, c_idx - 1 + ubicar_en_col, texto, centrado=False, negrita=negrita, fondo=self.uf.peach, colorido=self.uf.negro)

        self.auto_ajustar(worksheet)
        worksheet.sheet_view.showGridLines = False
        worksheet.sheet_view.zoomScale = 80
        if self.reporte_pie:
            self.escribir_footer(worksheet, len(df_parcial_tr), 4, self.nav)

        return worksheet

    def get_informe(self):
        '''creacion del informe'''
        workbook = Workbook()
        workbook.remove(workbook.active)
        if self.globales == 0:
            self.nav.nombre_analista = self.nav.nombre
            self.nav.nombre_supervisor = ''

        worksheet = workbook.create_sheet(title='Cadena de IVA', index=1)
        self.reporte_pie = True
        worksheet = self.get_hoja_cadena(worksheet)
        return workbook


class Provireporte(Pleyades.Abayo):
    '''Reporte de Providencias:'''

    xjd = ''
    nav = ''

    def __init__(self, db):
        '''constructor inicial'''
        super().__init__(db.uf)
        self.db = db
        self.uf = self.db.uf
        self.reporte_pie = False
        self.globales = 0
        self.df = botcito.DataFrame()

        self.procedencia = 'externa'
        if isinstance(self.xjd, str):
            self.procedencia = 'interna'
        match self.procedencia:
            case "externa":
                self.jd = self.xjd

            case "interna":
                self.jd = self.db.uf.pi
                self.nav = self.db.uf.navegante

        self._sql = RetencionesQ.NoCruzan(self.jd)

    def get_hoja_providencias(self, worksheet):
        prov = Providencias.Encontradas(self.db)
        _jd = self.uf.pi
        self._sql.jd.fecha_hoy = self.db.get_fecha_ymd()
        self.df, longitud_real, longitud_fisicas = prov.tratar_providencias(self._sql)
        worksheet.sheet_properties.tabColor = '1072BA'
        self.escribir_encabezado(worksheet, '', self._sql.jd)
        ubicar_en_fila, ubicar_en_col = 10, 4
        c_idx = 0
        columnas = ['EMISOR RUC', 'FECHA RETENCION','SECUENCIAL','AUTORIZACION','VALOR RETENCION','AÑO','MES','DOCUMENTO', 'FFPV']
        for columna in columnas:
            c_idx += 1
            self.escribir_celda(worksheet, ubicar_en_fila+2, ubicar_en_col + c_idx, columna, centrado=True, negrita=True, fondo=self.uf.sri, colorido=self.uf.blanco)

        ubicar_en_fila += 3
        df_parcial = self.df.reset_index()
        df_parcial = df_parcial.copy()
        for r_idx, fila in df_parcial.iterrows():
            c_idx = 1
            self.escribir_en_red(fila, worksheet, r_idx + ubicar_en_fila, c_idx + ubicar_en_col, "ruc_contrib_informan", "numero_ruc_emisor_pintar")
            c_idx += 1
            self.escribir_en_red(fila, worksheet, r_idx + ubicar_en_fila, c_idx + ubicar_en_col, "fecha_emi_retencion", "fecha_emision_pintar")
            c_idx += 1
            self.escribir_en_red(fila, worksheet, r_idx + ubicar_en_fila, c_idx + ubicar_en_col, "secuencial_retencion", "secuencial_pintar")     
            c_idx += 1
            self.escribir_en_red(fila, worksheet, r_idx + ubicar_en_fila, c_idx + ubicar_en_col, "autorizacion", "numero_autorizacion_pintar")
            c_idx += 1
            self.escribir_celda(worksheet, r_idx + ubicar_en_fila, c_idx + ubicar_en_col, fila["valor_retencion"], centrado=True, negrita=True, fondo=self.uf.blanco, colorido=self.uf.negro)
            c_idx += 1
            self.escribir_celda(worksheet, r_idx + ubicar_en_fila, c_idx + ubicar_en_col, fila["anio"], centrado=True, negrita=True, fondo=self.uf.blanco, colorido=self.uf.negro)
            c_idx += 1
            self.escribir_celda(worksheet, r_idx + ubicar_en_fila, c_idx + ubicar_en_col, fila["mes"], centrado=True, negrita=True, fondo=self.uf.blanco, colorido=self.uf.negro)

            if fila["documento"] != "RETENCION" and len(str(fila["documento"])) > 0:
                fondo = self.uf.rojo
            else:
                fondo = self.uf.blanco
            c_idx += 1
            self.escribir_celda(worksheet, r_idx + ubicar_en_fila, c_idx + ubicar_en_col, fila["documento"], centrado=True, negrita=True, fondo=fondo, colorido=self.uf.negro)

            if fila["ffpv"] != "si":
                fondo = self.uf.rojo
            else:
                fondo = self.uf.blanco
            c_idx += 1
            self.escribir_celda(worksheet, r_idx + ubicar_en_fila, c_idx + ubicar_en_col, fila["ffpv"], centrado=True, negrita=True, fondo=fondo, colorido=self.uf.negro)

        self.auto_ajustar(worksheet)
        worksheet.sheet_view.showGridLines = False
        worksheet.sheet_view.zoomScale = 80
        if self.reporte_pie:
            self.escribir_footer(worksheet, len(self.df.index), 6, self.nav)

        return worksheet

    def get_informe(self):
        workbook = Workbook()
        workbook.remove(workbook.active)
        if self.globales == 0:
            self.nav.nombre_analista = self.nav.nombre
            self.nav.nombre_supervisor = ''

        worksheet = workbook.create_sheet(title='Providencias', index=1)
        self.reporte_pie = True
        worksheet = self.get_hoja_providencias(worksheet)
        worksheet.title = f'Providencias ({len(self.df.index)})'

        return workbook


class Validacionesreporte(Pleyades.Abayo):
    '''
        Reporte de Providencias:

    '''
    xjd = ''
    nav = ''

    def __init__(self, db):
        '''constructor principal'''
        super().__init__(db.uf)
        self.db = db
        self.uf = self.db.uf
        self.hoja = ''
        self.df = []
        self.descripciones = []
        self.forma = 0
        self.columnas = []
        self.titulo_tabla = ''
        self.footer = 0
        self.df_nupy = {}
        self.globales = 0
        self.procedencia = 'externa'
        if isinstance(self.xjd, str):
            self.procedencia = 'interna'
        match self.procedencia:
            case "externa":
                self.jd = self.xjd

            case "interna":
                self.jd = self.db.uf.pi
                self.nav = self.db.uf.navegante

        self._sql = Reportes.Globales(self.jd)

    def escribir_informe_parecido(self):
        '''escribir informe '''
        self.hoja.sheet_properties.tabColor = '1072BA'
        self.escribir_encabezado(self.hoja, '', self._sql.jd)
        if len(self.titulo_tabla) > 0:
            c_idx = 2
            ubicar_en_fila, ubicar_en_col = 9, 4
            self.hoja.merge_cells('F11:G11')
            self.escribir_celda(self.hoja, ubicar_en_fila+2, ubicar_en_col + c_idx, self.titulo_tabla, centrado=True, negrita=True, fondo=self.uf.sri, colorido=self.uf.blanco)

        ubicar_en_fila, ubicar_en_col = 10, 4
        c_idx = 0
        for columna in self.columnas:
            c_idx += 1
            self.escribir_celda(self.hoja, ubicar_en_fila+2, ubicar_en_col + c_idx, columna, centrado=True, negrita=True, fondo=self.uf.sri, colorido=self.uf.blanco)
        ubicar_en_fila, ubicar_en_col = 10, 4
        ubicar_en_fila += 3
        df_parcial = self.df.copy()
        df_parcial = df_parcial.drop_duplicates()
        df_parcial["descripcion"] = self.descripciones

        c_idx = 0
        for r_idx, fila in df_parcial.iterrows():
            c_idx += 1
            self.escribir_celda(self.hoja, r_idx + ubicar_en_fila, c_idx + ubicar_en_col, fila["descripcion"], centrado=False, negrita=True, fondo=self.uf.blanco, colorido=self.uf.negro)
            c_idx += 1
            if self.forma == 1:
                self.escribir_celda(self.hoja, r_idx + ubicar_en_fila, c_idx + ubicar_en_col, fila["valor"], centrado=True, negrita=True, fondo=self.uf.blanco, colorido=self.uf.negro)

            if self.forma == 2:
                self.escribir_celda(self.hoja, r_idx + ubicar_en_fila, c_idx + ubicar_en_col, fila["valor1"], centrado=True, negrita=True, fondo=self.uf.blanco, colorido=self.uf.negro)
                c_idx += 1
                self.escribir_celda(self.hoja, r_idx + ubicar_en_fila, c_idx + ubicar_en_col, fila["valor2"], centrado=True, negrita=True, fondo=self.uf.blanco, colorido=self.uf.negro)

            c_idx = 0

        r_idx = 0
        c_idx = 0
        longitud_dataframe = len(df_parcial.index)

        if len(self.db.uf.pi.observaciones) > 0 and self.db.uf.pi.huesped == 'OBS':
            ubicar_en_fila, ubicar_en_col = 18 + longitud_dataframe, 4
            self.escribir_celda(self.hoja, r_idx + ubicar_en_fila, c_idx + ubicar_en_col, "OBSERVACIONES", centrado=True, negrita=True, fondo=self.uf.sri, colorido=self.uf.blanco)
            c_idx += 1
            self.escribir_celda(self.hoja, r_idx + ubicar_en_fila, c_idx + ubicar_en_col, self.db.uf.pi.observaciones, centrado=True, negrita=True, fondo=self.uf.blanco, colorido=self.uf.negro)
        if self.footer == 1:
            self.escribir_footer(self.hoja, longitud_dataframe, 5, self.nav)

        self.auto_ajustar(self.hoja)
        self.hoja.sheet_view.showGridLines = False
        self.hoja.sheet_view.zoomScale = 80

    def get_hoja_resumen_periodos(self, ws_periodos):
        '''resumen periodos'''
        ws_periodos.sheet_properties.tabColor = '1072BA'
        self.escribir_encabezado(ws_periodos, '', self._sql.jd)
        ubicar_en_fila, ubicar_en_col = 10, 4
        c_idx = 0
        columnas = ['FILA', 'AÑO', 'MES', 'TOTAL IMPUESTO A PAGAR', 'RETENCIONES A DEVOLVER', 'SALDOS']
        for columna in columnas:
            c_idx += 1
            self.escribir_celda(ws_periodos, ubicar_en_fila+2, ubicar_en_col + c_idx, columna, centrado=True, negrita=True, fondo=self.uf.sri, colorido=self.uf.blanco)
        ubicar_en_fila, ubicar_en_col = 10, 4
        ubicar_en_fila += 3
        c_idx = 0
        for r_idx, fila in self.df_nupy.iterrows():
            c_idx += 1
            self.escribir_celda(ws_periodos, r_idx + ubicar_en_fila, c_idx + ubicar_en_col, fila["fila"], centrado=True, negrita=True, fondo=self.uf.blanco, colorido=self.uf.negro)
            c_idx += 1
            self.escribir_celda(ws_periodos, r_idx + ubicar_en_fila, c_idx + ubicar_en_col, fila["anio"], centrado=True, negrita=True, fondo=self.uf.blanco, colorido=self.uf.negro)
            c_idx += 1
            self.uf.pi.mes = fila["mes"]
            self.escribir_celda(ws_periodos, r_idx + ubicar_en_fila, c_idx + ubicar_en_col, self.uf.get_mes_nombrado(), centrado=True, negrita=True, fondo=self.uf.blanco, colorido=self.uf.negro)
            c_idx += 1
            self.escribir_celda(ws_periodos, r_idx + ubicar_en_fila, c_idx + ubicar_en_col, fila["total_impuesto_a_pagar"], centrado=True, negrita=True, fondo=self.uf.blanco, colorido=self.uf.negro)
            c_idx += 1
            self.escribir_celda(ws_periodos, r_idx + ubicar_en_fila, c_idx + ubicar_en_col, fila["retenciones_a_devolver"], centrado=True, negrita=True, fondo=self.uf.blanco, colorido=self.uf.negro)
            c_idx += 1
            self.escribir_celda(ws_periodos, r_idx + ubicar_en_fila, c_idx + ubicar_en_col, fila["saldos"], centrado=True, negrita=True, fondo=self.uf.blanco, colorido=self.uf.negro)
            c_idx = 0

        # 1 pie de pagina
        self.footer = 0
        self.auto_ajustar(ws_periodos)
        ws_periodos.sheet_view.showGridLines = False
        ws_periodos.sheet_view.zoomScale = 80
        return ws_periodos

    def get_hoja_liquidacion(self, cade, _sql):
        '''hoja de liquidacion'''
        self.forma = 1
        self.df = cade.get_resumen_liquidacion(_sql).reset_index()
        self.descripciones = ['IMPUESTO CAUSADO',
                                'CRÉDITO TRIBUTARIO MES ACTUAL',
                                'IMPUESTO RESULTANTE',
                                'SALDO CREDITO TRIBUTARIO POR ADQUISICIONES E IMPORTACIONES - POR PAGO MEDIANTE USO DE MEDIOS ELECTRÓNICOS MES ANTERIOR',
                                'SALDO CREDITO TRIBUTARIO POR RETENCIONES EN LA FUENTE DE IVA QUE LE HAN SIDO EFECTUADAS APLICABLES MES ANTERIOR',
                                '( - ) RETENCIONES VERIFICADAS Y VÁLIDAS',
                                '(=) SALDO A FAVOR CONTRIBUYENTE',
                                'SALDO CREDITO TRIBUTARIO POR ADQUISICIONES E IMPORTACIONES - POR PAGO MEDIANTE USO DE MEDIOS ELECTRÓNICOS MES SIGUIENTE',
                                'SALDO CREDITO TRIBUTARIO POR RETENCIONES EN LA FUENTE DE IVA QUE LE HAN SIDO EFECTUADAS MES SIGUIENTE',
                                'RETENCIONES A FAVOR DEL CONTRIBUYENNTE',
                                '(-) VALOR RETENCIONES COMPENSADAS POSTERIORES AL PERIODO SOLICITADO',
                                'RETENCIONES A DEVOLVER',
                                'IMPUESTO A PAGAR']

        self.escribir_informe_parecido()

    def get_hoja_analizados(self, cade, _sql):
        '''hoja analizados'''
        self.df = cade.get_resumen_analizados(_sql).reset_index()
        self.descripciones = ['VALOR A RECONOCER POR RETENCIONES',
                                            'NO CONSTAN EN BASE (NB) ',
                                            'VALOR NO SUSTENTADO ',
                                            'VALOR NEGADO ',
                                            'DIFERENCIAS EN VALOR DECLARADO VS. LIBROS MAYORES',
                                            'CRÉDITO TRIBUTARIO PARA EL PRÓXIMO MES DE ACUERDO A ANÁLISIS ADQUISICIIONES',
                                            'CRÉDITO TRIBUTARIO PARA EL PRÓXIMO MES DE ACUERDO A ANÁLISIS RETENCIONES',
                                            'CT INCREMENTADO INJUSTIFICADAMENTE POR EL CONTRIBUYENTE ADQUISICIONES',
                                            'CT INCREMENTADO INJUSTIFICADAMENTE POR EL CONTRIBUYENTE RETENCIONES',
                                            'TOTAL VALORES ANÁLIZADOS']
        self.escribir_informe_parecido()

    def get_hoja_verificacion(self, cade, _sql):
        '''hoja de verificacion'''
        self.df = cade.get_resumen_verificados(_sql).reset_index()
        self.forma = 2
        self.descripciones = ['(+) TOTAL VALORES ANÁLIZADOS',
                                            '(-) CREDITO TRIBUTARIO PARA EL PROXIMO MES DECLARADO EN EL ULTIMO MES ANALIZADO (ADQUISICIONES E IMPORTACIONES)',
                                            '(-) CREDITO TRIBUTARIO PARA EL PROXIMO MES DECLARADO EN EL ULTIMO MES ANALIZADO (RETENCIONES)',
                                            'DIFERENCIA EN CADENA A REVISAR']
        self.columnas = ['DETALLE','VALOR','OBSERVACIONES']
        self.escribir_informe_parecido()

    def get_hoja_resultados(self, cade, _sql):
        '''hoja de resultado'''
        self.df = cade.get_resumen_resultados(_sql).reset_index()
        _sql.jd.observaciones = cade.get_resumen_resultados_obs(_sql)
        self.forma = 2
        self.titulo_tabla = 'CRÉDITO TRIBUTARIO PARA EL PROXIMO MES DECLARADO EN EL ÚLTIMO MES ANALIZADO'
        self.descripciones = ['VALOR DECLARADO',
                                        'CT INCREMENTADO INJUSTIFICADAMENTE POR EL CONTRIBUYENTE',
                                        'RETENCIONES A DEVOLVER',
                                        'NEGADOS / NO CONSTAN BASE/ NO SUSTENTADOS / DIF. MAYORES',
                                        'TOTAL VALORES ANALIZADOS',
                                        'VALOR FINAL DE LA CADENA',
                                        'DIFERENCIAS A REVISAR']
        self.columnas = ['DETALLE', 'ADQUISICIONES E IMPORTACIONES', 'RETENCIONES']
        self.footer = 1
        _sql.jd.huesped = 'OBS'
        self.escribir_informe_parecido()
        _sql.jd.huesped = ''

    def get_informe(self):
        '''get informe'''
        cade = Cadena.Iva(self.db)
        if self.globales == 0:
            self.nav.nombre_analista = self.nav.nombre
            self.nav.nombre_supervisor = ''
        self.df_nupy = cade.get_resumen_periodos(self._sql).reset_index()
        workbook = Workbook()
        workbook.remove(workbook.active)
        # resumen periodos
        ws_periodos = workbook.create_sheet(title=f'Resumen Periodos ({len(self.df_nupy.index)})', index=1)
        ws_periodos = self.get_hoja_resumen_periodos(ws_periodos)
        self.columnas = ['DETALLE', 'VALOR']
        # RESUMEN VALORES ANALIZADOS
        self.hoja = workbook.create_sheet(title='Cuadro Liquidacion', index=3)
        self.get_hoja_liquidacion(cade, self._sql)
        # RESUMEN VALORES ANALIZADOS
        self.hoja = workbook.create_sheet(title='Valores Analizados', index=3)
        self.get_hoja_analizados(cade, self._sql)
        # 1 VERIFICACIÓN DE RESULTADOS (CUADRE DE CADENA)
        self.hoja = workbook.create_sheet(title='Verificacion', index=4)
        self.get_hoja_verificacion(cade, self._sql)        
        # 2 VERIFICACIÓN DE RESULTADOS (CUADRE DE CADENA)
        self.hoja = workbook.create_sheet(title='Resultados', index=4)
        self.get_hoja_resultados(cade, self._sql)

        return workbook


class Resumen_pers_repo(Pleyades.Abayo):
    '''
        Reporte de Resumen Periodos:
    '''

    xjd = ''
    nav = ''

    def __init__(self, db):
        '''constructor principal'''
        super().__init__(db.uf)
        self.db = db
        self.uf = self.db.uf
        self.reporte_pie = False
        self.globales = 0
        self.df = botcito.DataFrame()

        self.procedencia = 'externa'
        if isinstance(self.xjd, str):
            self.procedencia = 'interna'
        match self.procedencia:
            case "externa":
                self.jd = self.xjd

            case "interna":
                self.jd = self.db.uf.pi
                self.nav = self.db.uf.navegante
        self._sql = Reportes.Globales(self.jd)

    def get_hoja_periodos(self, worksheet):
        '''hoja de periodos'''
        info = Informe.Revision(self.db)
        self.df = info.get_mayoreo(self._sql)

        df_parcial = self.df.reset_index()
        df_parcial = df_parcial.copy()
        worksheet.sheet_properties.tabColor = '1072BA'
        self.escribir_encabezado(worksheet, '', self._sql.jd)
        ubicar_en_fila, ubicar_en_col = 10, 4
        c_idx = 0
        columnas = ['AÑO', 'MES', 'VALOR_DECLARADO', 'MAYORES', 'DIFERENCIA', 'TOTAL', 'NO CONSTA EN BASE', 'NO SUSTENTADO', 'NEGADO', 'ACEPTADO', 'A LA CADENA']
        for columna in columnas:
            c_idx += 1
            self.escribir_celda(worksheet, ubicar_en_fila+2, ubicar_en_col + c_idx, columna, centrado=True, negrita=True, fondo=self.uf.sri, colorido=self.uf.blanco)

        ubicar_en_fila += 3

        df_parcial["anio"] = df_parcial["anio"].astype(int) 
        df_parcial["mes"] = df_parcial["mes"].astype(int)
        df_parcial.loc['total'] = df_parcial.sum()
        df_parcial.drop(["index"],inplace=True, axis=1)
        df_parcial["anio"] = df_parcial["anio"].astype(str) 
        df_parcial["mes"] = df_parcial["mes"].astype(str) 
        df_parcial["numero_adhesivo"] = df_parcial["numero_adhesivo"].astype(str)
        df_parcial.loc[df_parcial.index[-1], 'anio'] = 'TOTAL'
        df_parcial.loc[df_parcial.index[-1], 'mes'] = ''
        df_parcial.loc[df_parcial.index[-1], 'numero_adhesivo'] = ''
        r_idx = -1
        longitud = len(df_parcial.index)
        for r_idx_1, fila in df_parcial.iterrows():
            r_idx += 1
            c_idx = 1
            fondo_eleccion = self.uf.blanco
            if longitud == r_idx+1:
                fondo_eleccion = self.uf.lavender
            self.escribir_celda(worksheet, r_idx + ubicar_en_fila, c_idx + ubicar_en_col, str(fila["anio"]), centrado=True, negrita=True, fondo=fondo_eleccion, colorido=self.uf.negro)
            c_idx += 1
            self.uf.pi.mes = str(fila["mes"])
            mes = self.uf.get_mes_nombrado()
            self.escribir_celda(worksheet, r_idx + ubicar_en_fila, c_idx + ubicar_en_col, mes, centrado=True, negrita=True, fondo=fondo_eleccion, colorido=self.uf.negro)
            c_idx += 1
            self.escribir_celda(worksheet, r_idx + ubicar_en_fila, c_idx + ubicar_en_col, fila["retenciones_fuente_iva"], centrado=True, negrita=True, fondo=fondo_eleccion, colorido=self.uf.negro)
            c_idx += 1
            self.escribir_celda(worksheet, r_idx + ubicar_en_fila, c_idx + ubicar_en_col, fila["mayores"], centrado=True, negrita=True, fondo=fondo_eleccion, colorido=self.uf.negro)
            c_idx += 1
            self.escribir_celda(worksheet, r_idx + ubicar_en_fila, c_idx + ubicar_en_col, fila["diferencia_actualizar"], centrado=True, negrita=True, fondo=fondo_eleccion, colorido=self.uf.negro)
            c_idx += 1
            self.escribir_celda(worksheet, r_idx + ubicar_en_fila, c_idx + ubicar_en_col, fila["ingresados"], centrado=True, negrita=True, fondo=fondo_eleccion, colorido=self.uf.negro)
            c_idx += 1
            self.escribir_celda(worksheet, r_idx + ubicar_en_fila, c_idx + ubicar_en_col, fila["no_consta_base"], centrado=True, negrita=True, fondo=fondo_eleccion, colorido=self.uf.negro)
            c_idx += 1
            self.escribir_celda(worksheet, r_idx + ubicar_en_fila, c_idx + ubicar_en_col, fila["nocruzan"], centrado=True, negrita=True, fondo=fondo_eleccion, colorido=self.uf.negro)
            c_idx += 1
            self.escribir_celda(worksheet, r_idx + ubicar_en_fila, c_idx + ubicar_en_col, fila["negados_dups"], centrado=True, negrita=True, fondo=fondo_eleccion, colorido=self.uf.negro)
            c_idx += 1
            self.escribir_celda(worksheet, r_idx + ubicar_en_fila, c_idx + ubicar_en_col, fila["aceptados_cadena"], centrado=True, negrita=True, fondo=fondo_eleccion, colorido=self.uf.negro)
            c_idx += 1
            self.escribir_celda(worksheet, r_idx + ubicar_en_fila, c_idx + ubicar_en_col, fila["aceptados_cadena"], centrado=True, negrita=True, fondo=fondo_eleccion, colorido=self.uf.negro)

        self.auto_ajustar(worksheet)
        worksheet.sheet_view.showGridLines = False
        worksheet.sheet_view.zoomScale = 80

        if self.reporte_pie:
            self.escribir_footer(worksheet, len(df_parcial.index), 6, self.nav)

        return worksheet

    def get_informe(self):
        '''get informe'''
        workbook = Workbook()
        workbook.remove(workbook.active)
        worksheet = workbook.create_sheet(title='Resumen Periodos', index=1)
        self.reporte_pie = True
        if self.globales == 0:
            self.db.uf.navegante.nombre_analista = self.db.uf.navegante.nombre
            self.db.uf.navegante.nombre_supervisor = ''
        worksheet = self.get_hoja_periodos(worksheet)
        worksheet.title = f'Resumen Periodos ({len(self.df.index)})'
        return workbook


class Globalreporte(Pleyades.Abayo):
    '''
        Reporte de Global:
    '''
    def __init__(self, db):
        '''constructor principal'''
        super().__init__(db.uf)
        self.db = db
        self.uf = self.db.uf
        self.cn = Consultas.Papel(db)
        self._jd_espejo = self.db.uf.pi
        self._sql = Reportes.Globales(self._jd_espejo)

    def get_sql_razon_social(self, contri):
        '''razon social'''
        if len(contri) == 10:
            contri = self.db.uf.costelo(contri)
        return self.cn.get_sql_razon_social_his(contri)

    def get_informe(self):
        '''informe global'''
        xjd = self._jd_espejo
        nav = self.db.uf.navegante
        df_memoria = \
            self.db.get_vector(self._sql.get_sql_resumen_tramites_id())
        if not df_memoria.empty:
            xjd.tramite = df_memoria["numero_tramite"].iloc[0]
            xjd.contri = df_memoria["contri"].iloc[0]
            xjd.usuario = df_memoria["usuario"].iloc[0]
            xjd.periodo_inicial = df_memoria["periodo_inicial"].iloc[0]
            xjd.periodo_final = df_memoria["periodo_final"].iloc[0]
            xjd.periodo_inicial_org = df_memoria["periodo_inicial"].iloc[0]
            xjd.periodo_final_org = df_memoria["periodo_final"].iloc[0]
            xjd.procedencia = "externa"
            nav.nombre_analista = df_memoria["nombre_analista"].iloc[0]
            nav.nombre_supervisor = df_memoria["nombre_supervisor"].iloc[0]
        xjd.razon_social = \
            self.db.get_scalar(self.get_sql_razon_social(xjd.contri))

        workbook = Workbook()
        workbook.remove(workbook.active)
        worksheet = workbook.create_sheet(title='IR', index=1)

        seccion_ir = InformeRevreporte(self.db)
        seccion_ir.xjd = xjd
        seccion_ir.nav = nav
        seccion_ir.globales = 1
        seccion_ir.reporte_pie = False
        worksheet = seccion_ir.get_hoja_cruces(worksheet)
        worksheet.title = f'Cruce Retenciones ({len(seccion_ir.df.index)})'

        # NUEVO CADENA EARLIER 684
        worksheet = \
            workbook.create_sheet(title='Reporte Analisis Previo Cadena',
                                  index=2)
        seccion_ire = InformeCadenaEarlier_reporte(self.db)
        seccion_ire.xjd = xjd
        seccion_ire.nav = nav
        seccion_ire.globales = 1
        seccion_ire.reporte_pie = False
        worksheet = seccion_ire.get_hoja_analisis(worksheet)
        worksheet.title = f'Analisis ({len(seccion_ire.dff.index)})'
        # FIN NUEVO

        worksheet = workbook.create_sheet(title='Providencias', index=3)
        seccion_ir = Provireporte(self.db)
        seccion_ir.xjd = xjd
        seccion_ir.nav = nav
        seccion_ir.globales = 1
        seccion_ir.reporte_pie = False
        worksheet = seccion_ir.get_hoja_providencias(worksheet)
        worksheet.title = f'Providencias ({len(seccion_ir.df.index)})'

        worksheet = workbook.create_sheet(title='Resumen Periodos', index=4)
        seccion_ir = Resumen_pers_repo(self.db)
        seccion_ir.xjd = xjd
        seccion_ir.nav = nav
        seccion_ir.globales = 1
        seccion_ir.reporte_pie = False
        worksheet = seccion_ir.get_hoja_periodos(worksheet)
        worksheet.title = f'Resumen Periodos ({len(seccion_ir.df.index)})'

        worksheet = workbook.create_sheet(title='Cadena de IVA', index=5)
        seccion_ir = Cadenareporte(self.db)
        seccion_ir.xjd = xjd
        seccion_ir.nav = nav
        seccion_ir.globales = 1
        seccion_ir.reporte_pie = False
        worksheet = seccion_ir.get_hoja_cadena(worksheet)
        worksheet.title = f'Cadena de IVA ({len(seccion_ir.df.index)})'

        # nuevo compensacion a futuro 723
        worksheet = workbook.create_sheet(title='Futuro', index=6)
        seccion_ir = InformeFuturas_reporte(self.db)
        seccion_ir.xjd = xjd
        seccion_ir.nav = nav
        seccion_ir.globales = 1
        seccion_ir.reporte_pie = False
        worksheet = seccion_ir.get_hoja_cruces(worksheet)
        worksheet.title = f'Futuro ({len(seccion_ir.df.index)})'
        #  fin compensa a futuro

        #  resumen periodos
        seccion_ir = Validacionesreporte(self.db)
        seccion_ir.xjd = xjd
        seccion_ir.nav = nav
        seccion_ir.globales = 1

        cade = Cadena.Iva(self.db)
        seccion_ir.df_nupy = \
            cade.get_resumen_periodos(seccion_ir._sql).reset_index()

        itt = len(seccion_ir.df_nupy.index)
        ws_periodos = \
            workbook.create_sheet(title=f'Resumen Periodos Con  ({itt})',
                                  index=7)
        ws_periodos = seccion_ir.get_hoja_resumen_periodos(ws_periodos)

        seccion_ir.columnas = ['DETALLE', 'VALOR']
        # RESUMEN VALORES ANALIZADOS
        seccion_ir.hoja = workbook.create_sheet(title='Cuadro Liquidacion',
                                                index=8)
        seccion_ir.get_hoja_liquidacion(cade, self._sql)

        # RESUMEN VALORES ANALIZADOS
        seccion_ir.hoja = workbook.create_sheet(title='Valores Analizados',
                                                index=9)
        seccion_ir.get_hoja_analizados(cade, self._sql)

        # VERIFICACIÓN DE RESULTADOS (CUADRE DE CADENA)
        seccion_ir.hoja = workbook.create_sheet(title='Verificacion', index=10)
        seccion_ir.get_hoja_verificacion(cade, self._sql)

        # VERIFICACIÓN DE RESULTADOS (CUADRE DE CADENA)
        seccion_ir.hoja = workbook.create_sheet(title='Resultados', index=11)
        seccion_ir.get_hoja_resultados(cade, self._sql)

        self.db.uf.pi.tramite = self.db.uf.espejo.tramite
        self.db.uf.pi.usuario = self.db.uf.espejo.usuario
        self.db.uf.pi.procedencia = 'interna'
        return workbook


class InformeDeclas_reporte(Pleyades.Abayo):
    ''' Reporte de Revision:     '''
    def __init__(self, db):
        '''constructor principal'''
        super().__init__(db.uf)
        self.db = db
        self.uf = self.db.uf
        self.reporte_pie = False
        self.globales = 0
        self.df = botcito.DataFrame()
        self._sql = Reportes.Globales(self.uf.pi)
        self.nav = self.uf.navegante

    def get_hoja_cruces(self, worksheet):
        '''hoja de cruces'''
        self.db.uf.pi.ufuf = -12
        self._sql.jd.ufuf = -12
        info = Fotones.Reemplazantes(self.db)
        self.df = info.volcado(self._sql)

        df = self.df.reset_index()
        df.drop(columns=["index"], axis=1, inplace=True)
        df_pivot = df.T.reset_index()
        ix = 0
        for col in df_pivot.columns:
            if ix > 0:
                mes = df_pivot.iloc[1, ix]
                df_pivot = df_pivot.rename(columns={col: meses(mes)})
            ix += 1

        df_pivot = df_pivot.rename(columns={"index": "CONCEPTO"})
        df_pivot.replace({'numero_adhesivo': 'ADHESIVO'}, inplace=True)
        df_pivot.replace({'fecha_recepcion': 'FECHA RECEPCION'}, inplace=True)
        df_pivot.replace({'codigo_impuesto': 'CODIGO DE IMPUESTO'}, inplace=True)
        df_pivot.replace({'anio_fiscal': 'AÑO FISCAL'}, inplace=True)
        df_pivot.replace({'mes_fiscal': 'MES FISCAL'}, inplace=True)
        df_pivot.replace({'tot_imp_vnl_mac_iaf_1260': 'TOTAL IMPUESTO VENTAS NETAS A LIQUIDAR MES ACTUAL INCLUYE ACTIVOS FIJOS'}, inplace=True)
        df_pivot.replace({'crt_acu_fap_2130': 'CREDITO TRIBUTARIO DE ACUERDO A FACTOR DE PROPORCIONALIDAD'}, inplace=True)
        df_pivot.replace({'impuesto_causado_2140': 'IMPUESTO CAUSADO'}, inplace=True)
        df_pivot.replace({'credito_tributario_mac_2150': 'CREDITO TRIBUTARIO MES ACTUAL'}, inplace=True)
        df_pivot.replace({'saldo_crt_cle_man_2160': 'SALDO CREDITO TRIBUTARIO POR COMPRAS LOCALES E IMPORTACIONES MES ANTERIOR'}, inplace=True)
        df_pivot.replace({'saldo_crt_rfu_man_2170': 'SALDO CREDITO TRIBUTARIO POR RETENCIONES EN LA FUENTE DE IVA QUE LE HAN SIDO EFECTUADAS MES ANTERIOR'}, inplace=True)
        df_pivot.replace({'rfu_mes_actual_2200': 'RETENCIONES EN LA FUENTE DE IVA QUE LE HAN SIDO EFECTUADAS MES ACTUAL'}, inplace=True)
        df_pivot.replace({'saldo_crt_clo_ipr_msi_2220': 'SALDO CREDITO TRIBUTARIO POR COMPRAS LOCALES E IMPORTACIONES MES SIGUIENTE'}, inplace=True)
        df_pivot.replace({'saldo_crt_rfu_msi_2230': 'SALDO CREDITO TRIBUTARIO POR RETENCIONES EN LA FUENTE DE IVA QUE LE HAN SIDO EFECTUADAS MES SIGUIENTE'}, inplace=True)
        df_pivot.replace({'aju_idr_pc_ipt_crt_mac_2210': 'AJUSTE POR IVA DEVUELTO E IVA RECHAZADO (POR CONCEPTO DE DEVOLUCIONES DE IVA), AJUSTE DE IVA POR PROCESOS DE CONTROL Y OTROS (ADQUISICIONES EN IMPORTACIONES), IMPUTABLES AL CREDITO TRIBUTARIO'}, inplace=True)
        df_pivot.replace({'aju_idr_pc_ipt_crt_mac_rf_2212': 'AJUSTE POR IVA DEVUELTO E IVA RECHAZADO, AJUSTE DE IVA POR PROCESOS DE CONTROL Y OTROS (POR CONCEPTO RETENCIONES EN LA FUENTE DE IVA), IMPUTABLES AL CRÉDITO TRIBUTARIO'}, inplace=True)
        df_pivot.replace({'subtotal_apa_aip_2250': 'SUBTOTAL A PAGAR ANTES DE IVA PRESUNTIVO'}, inplace=True)
        df_pivot.replace({'tot_imp_apa_percepcion_2270': 'TOTAL IMPUESTO A PAGAR POR PERCEPCION'}, inplace=True)
        df_pivot.replace({'tot_iva_retenido_2550': 'TOTAL IMPUESTO RETENIDO'}, inplace=True)
        df_pivot.replace({'total_pagado_2640': 'TOTAL PAGADO'}, inplace=True)
        df_pivot.replace({'vlb_eaf_tdc_450': 'VENTAS LOCALES (EXCLUYE ACTIVOS FIJOS) GRAVADAS TARIFA DIFERENTE DE CERO '}, inplace=True)
        df_pivot.replace({'tot_vln_eaf_tdc_480': 'TOTAL VENTAS LOCALES NETAS EXCLUYE ACTIVOS FIJOS TARIFA DIFERENTE DE CERO'}, inplace=True)
        df_pivot.replace({'total_pagado': 'TOTAL PAGADO'}, inplace=True)

        self.df = df_pivot

        rows = dataframe_to_rows(self.df)
        worksheet.sheet_properties.tabColor = '1072BA'
        worksheet.merge_cells('E4:I4')
        worksheet.merge_cells('E5:I5')
        worksheet.merge_cells('F6:I6')
        worksheet.merge_cells('F7:I7')
        worksheet.merge_cells('F8:I8')
        if self.globales == 0:
            self.nav.nombre_analista = self.nav.nombre
            self.nav.nombre_supervisor = ''

        self.escribir_encabezado(worksheet, '', self._sql.jd)

        ubicar_en_fila, ubicar_en_col = 10, 4
        for r_idx, row in enumerate(rows, 1):
            for c_idx, valor in enumerate(row, 1):
                if (c_idx < 2):
                    continue
                if r_idx == 1:
                    self.escribir_celda(worksheet, r_idx+ubicar_en_fila, c_idx-1+ubicar_en_col, valor, centrado=True, negrita=True, fondo=self.uf.sri, colorido=self.uf.blanco)
                else:
                    self.escribir_celda(worksheet, r_idx - 1 + ubicar_en_fila, c_idx-1+ubicar_en_col, valor, centrado=False, negrita=False, fondo=self.uf.peach, colorido=self.uf.negro)

        self.auto_ajustar(worksheet)
        worksheet.sheet_view.showGridLines = False
        worksheet.sheet_view.zoomScale = 80

        if self.reporte_pie:
            self.escribir_footer(worksheet, len(self.df.index), 6, self.nav)

        return worksheet

    def get_informe(self):
        '''informe'''
        self._sql.jd.procedencia = 'interna'
        workbook = Workbook()
        workbook.remove(workbook.active)
        worksheet = workbook.create_sheet(title='Declaraciones', index=1)
        self.reporte_pie = True
        worksheet = self.get_hoja_cruces(worksheet)
        worksheet.title = f'Declaraciones ({len(self.df.index)})'
        return workbook


class InformeFuturas_reporte(Pleyades.Abayo):
    ''' Reporte de Compensacion a Futura:'''
    xjd: str = ''

    def __init__(self, db):
        '''constructor principal'''
        super().__init__(db.uf)
        self.db = db
        self.uf = self.db.uf
        self.reporte_pie = False
        self.globales = 0
        self.df = botcito.DataFrame()

        self.nav = self.uf.navegante
        self.procedencia = 'externa'
        if isinstance(self.xjd, str):
            self.procedencia = 'interna'
        match self.procedencia:
            case "externa":
                self.jd = self.xjd

            case "interna":
                self.jd = self.db.uf.pi
                self.nav = self.db.uf.navegante
        self._sql = Reportes.Globales(self.uf.pi)

    def get_hoja_cruces(self, worksheet):
        '''hoja de cruces'''
        self.db.uf.pi.ufuf = -13
        self._sql.jd.ufuf = -13
        info = Fotones.Reemplazantes(self.db)
        self.df = info.volcado(self._sql)
        df = self.df.reset_index()
        df.drop(columns=["index"], axis=1, inplace=True)
        df["anio_fiscal"] = df["anio_fiscal"].astype(str)
        df["mes_fiscal"] = df["mes_fiscal"].astype(str)
        df_pivot = df.T.reset_index()
        ix = 0
        for col in df_pivot.columns:
            if ix > 0:
                mes = df_pivot.iloc[1, ix]
                df_pivot = df_pivot.rename(columns={col: meses(mes)})
            ix += 1

        df_pivot = df_pivot.rename(columns={"index": "CONCEPTO"})
        df_pivot.replace({'anio_fiscal': 'AÑO FISCAL'}, inplace=True)
        df_pivot.replace({'mes_fiscal': 'MES FISCAL'}, inplace=True)
        df_pivot.replace({'camino': 'INFERENCIA'}, inplace=True)
        df_pivot.replace({'fecha_recepcion': 'FECHA RECEPCION'}, inplace=True)
        df_pivot.replace({'sustitutiva_original': 'SUSTITUTIVA ORIGINAL'}, inplace=True)
        df_pivot.replace({'numero_adhesivo': 'NUMERO ADHESIVO'}, inplace=True)
        df_pivot.replace({'sct_credito_mes_anterior_rca_adq_ret': 'SALDO CREDITO TRIBUTARIO MES ANTERIOR RCA ADQUISICION RETENCION'}, inplace=True)
        df_pivot.replace({'ajuste_x_adquisiciones': 'AJUSTE X ADQUISICIONES'}, inplace=True)
        df_pivot.replace({'saldo_de_ct_mes_anterior': 'SALDO DE CT MES ANTERIOR'}, inplace=True)
        df_pivot.replace({'impuesto_causado': 'IMPUESTO CAUSADO'}, inplace=True)
        df_pivot.replace({'ct_mes_actual': 'CT MES ACTUAL'}, inplace=True)
        df_pivot.replace({'retenciones_fuente_iva': 'RETENCIONES FUENTE IVA'}, inplace=True)
        df_pivot.replace({'tot_impuesto_pagar_x_percepcion': 'TOTAL IMPUESTO PAGAR X PERCEPCION'}, inplace=True)
        df_pivot.replace({'compensa_futuro_reco_sol_atendida': 'COMPENSA FUTURA RECO SOL ATENDIDA'}, inplace=True)
        df_pivot.replace({'saldo_cred_resulta_next_mes': 'SALDO CREDITO RESULTA SIGUIENTE MES'}, inplace=True)
        df_pivot.replace({'impuesto_pagar_resulta_mes': 'IMPUESTO A PAGAR RESULTA MES'}, inplace=True)
        self.df = df_pivot
        rows = dataframe_to_rows(self.df)
        worksheet.sheet_properties.tabColor = '1072BA'
        # 1 worksheet.freeze_panes = 'L2'
        worksheet.merge_cells('E4:I4')
        worksheet.merge_cells('E5:I5')
        worksheet.merge_cells('F6:I6')
        worksheet.merge_cells('F7:I7')
        worksheet.merge_cells('F8:I8')
        if self.globales == 0:
            self.db.uf.navegante.nombre_analista = self.db.uf.navegante.nombre
            self.db.uf.navegante.nombre_supervisor = ''
        self.escribir_encabezado(worksheet, '', self._sql.jd)

        ubicar_en_fila, ubicar_en_col = 10, 4
        for r_idx, row in enumerate(rows, 1):
            for c_idx, valor in enumerate(row, 1):
                if (c_idx < 2):
                    continue
                if r_idx == 1:
                    self.escribir_celda(worksheet, r_idx+ubicar_en_fila, c_idx - 1 + ubicar_en_col, valor, centrado=True, negrita=True, fondo=self.uf.sri, colorido=self.uf.blanco)
                else:
                    self.escribir_celda(worksheet, r_idx - 1 + ubicar_en_fila, c_idx - 1 + ubicar_en_col, valor, centrado=False, negrita=False, fondo=self.uf.peach, colorido=self.uf.negro)

        self.auto_ajustar(worksheet)
        worksheet.sheet_view.showGridLines = False
        worksheet.sheet_view.zoomScale = 80
        if self.reporte_pie:
            self.escribir_footer(worksheet, len(self.df.index), 6, self.nav)
        return worksheet

    def get_informe(self):
        '''informe principal'''
        self._sql.jd.procedencia = 'interna'
        workbook = Workbook()
        workbook.remove(workbook.active)
        worksheet = workbook.create_sheet(title='Comp. Futura', index=1)
        self.reporte_pie = True
        worksheet = self.get_hoja_cruces(worksheet)
        worksheet.title = f'Compensación Futura ({len(self.df.index)})'
        return workbook


class InformeCadenaEarlier_reporte(Pleyades.Abayo):
    '''Reporte de Cadena Fase 1:  '''
    xjd = ''

    def __init__(self, db):
        '''constructor principal'''
        super().__init__(db.uf)
        self.db = db
        self.uf = self.db.uf
        self.reporte_pie = False
        self.globales = 0
        self.dff = botcito.DataFrame()
        self.nav = self.uf.navegante

        self.procedencia = 'externa'
        if isinstance(self.xjd, str):
            self.procedencia = 'interna'
        match self.procedencia:
            case "externa":
                self.jd = self.xjd

            case "interna":
                self.jd = self.db.uf.pi
                self.nav = self.db.uf.navegante

        self._sql = Reportes.Globales(self.uf.pi)

    def get_ivot(self, df_pivot):
        '''pivot'''
        df_pivot = df_pivot.rename(columns={"index": "CONCEPTO"})
        df_pivot = df_pivot.rename(columns={'anio': 'AÑO FISCAL'})
        df_pivot = df_pivot.rename(columns={'mes': 'MES FISCAL'})
        df_pivot = df_pivot.rename(columns={'camino': 'INFERENCIA'})
        df_pivot = df_pivot.rename(columns={'fecha_recepcion': 'FECHA RECEPCION'})
        df_pivot = df_pivot.rename(columns={'numero_adhesivo': 'NUMERO ADHESIVO'})
        df_pivot = df_pivot.rename(columns={'sct_retenciones_mesanterior':
            'SALDO CREDITO TRIBUTARIO RETENCIONES MES ANTERIOR'})
        return df_pivot

    def get_hoja_analisis(self, worksheet_cea):
        '''hoja de analisis'''
        self.db.uf.pi.ufuf = -14
        self._sql.jd.ufuf = -14
        info = Fotones.Reemplazantes(self.db)
        self.dff = info.volcado(self._sql)
        df = self.dff.reset_index()
        df.drop(columns=["index"], axis=1, inplace=True)
        df.anio_fiscal = df.anio_fiscal.astype(str)
        df.mes_fiscal = df.mes_fiscal.astype(str)
        df.fecha_recepcion = df.fecha_recepcion.astype(str)
        df.codigo_impuesto = df.codigo_impuesto.astype(str)
        df_pivot = df.T.reset_index()
        ix = 0
        for col in df_pivot.columns:
            if ix > 0:
                mes = df_pivot.iloc[1, ix]
                df_pivot = df_pivot.rename(columns={col: meses(mes)})
            ix += 1

        df_pivot = df_pivot.rename(columns={"index": "CONCEPTO"})
        df_pivot.CONCEPTO = df_pivot.CONCEPTO.str.strip()
        df_pivot.CONCEPTO = df_pivot.CONCEPTO.astype("category")
        # 1 'total_pagado', 'total_impuesto_a_pagar_2610',

        print(f"df_pivot PRE {df_pivot} \n ")

        categorias = ['anio_fiscal', 'mes_fiscal', 'camino',
                      'fecha_recepcion', 'tipo_declaracion',
                      'numero_adhesivo', 'sct_adquisicion_mesanterior',
                      'sct_retenciones_mesanterior', 'ajuste_x_adquisiciones',
                      'ajuste_x_retenciones',  'sct_mes_anterior',
                      'sct_mesanterior_retenciones',
                      'total_impuestos_mes_actual',
                      'ct_factor_proporcionalidad',
                      'impuesto_causado', 'ct_mes_actual',
                      'retenciones_fuente_iva', 'sct_x_adquisiciones',
                      'sct_x_retenciones',
                      'tot_impuesto_pagar_x_percepcion',
                      'diferencia_arr_ct', 'diferencia_x_ct',
                      'diferencia_adquisiciones', 'diferencia_retenciones',
                      'totales', 'codigo_impuesto',
                      'calculo_ct_adq', 'calculo_ct_ret'
                      ]

        df_pivot.CONCEPTO = botcito.Categorical(df_pivot.CONCEPTO,
                                                categories=categorias)

        df_pivot = df_pivot.sort_values('CONCEPTO')
        df_pivot = df_pivot.reset_index()
        df_pivot.drop("index", axis=1, inplace=True)

        df_pivot = self.get_ivot(df_pivot=df_pivot)

        df_pivot.CONCEPTO = df_pivot.CONCEPTO.astype(str)
        df_pivot.replace({'anio_fiscal': 'AÑO FISCAL'}, inplace=True)
        df_pivot.replace({'mes_fiscal': 'MES FISCAL'}, inplace=True)
        df_pivot.replace({'camino': 'CAMINO'}, inplace=True)
        df_pivot.replace({'fecha_recepcion': 'FECHA RECEPCION'}, inplace=True)
        df_pivot.replace({'tipo_declaracion': 'TIPO DE DECLARACION'}, inplace=True)
        df_pivot.replace({'numero_adhesivo': 'NUMERO ADHESIVO'}, inplace=True)
        df_pivot.replace({'sct_adquisicion_mesanterior': 'SALDO CREDITO TRIBUTARIO POR ADQUISICIONES MES ANTERIOR'}, inplace=True)
        df_pivot.replace({'sct_retenciones_mesanterior': 'SALDO CREDITO TRIBUTARIO POR RETENCIONES MES ANTERIOR'}, inplace=True)
        df_pivot.replace({'ajuste_x_adquisiciones': 'AJUSTE POR IVA DEVUELTO POR ADQUISICIONES'}, inplace=True)
        df_pivot.replace({'ajuste_x_retenciones': 'AJUSTE POR IVA DEVUELTO POR RETENCIONES'}, inplace=True)
        df_pivot.replace({'sct_mes_anterior': 'SALDO CREDITO TRIBUTARIO POR ADQUISICIONES MES ANTERIOR'}, inplace=True)
        df_pivot.replace({'sct_mesanterior_retenciones': 'SALDO CREDITO TRIBUTARIO POR RETENCIONES MES ANTERIOR'}, inplace=True)          
        df_pivot.replace({'total_impuestos_mes_actual': 'TOTAL IMPUESTO VENTAS A LIQUIDAR MES ACTUAL'}, inplace=True)
        df_pivot.replace({'ct_factor_proporcionalidad': 'CREDITO TRIBUTARIO APLICABLE EN ESTE PERIODO (FACTOR DE PROPORCIONALIDAD)'}, inplace=True)
        df_pivot.replace({'impuesto_causado': 'IMPUESTO CAUSADO'}, inplace=True)
        df_pivot.replace({'ct_mes_actual': 'CREDITO TRIBUTARIO MES ACTUAL'}, inplace=True)
        df_pivot.replace({'retenciones_fuente_iva': '(-) RETENCIONES EN LA FUENTE DE IVA QUE LE HAN SIDO EFECTUADAS'}, inplace=True)
        df_pivot.replace({'sct_x_adquisiciones': 'SALDO CREDITO TRIBUTARIO POR ADQUISICIONES MES SIGUIENTE'}, inplace=True)
        df_pivot.replace({'sct_x_retenciones': 'SALDO CREDITO TRIBUTARIO POR RETENCIONES MES SIGUIENTE'}, inplace=True)
        df_pivot.replace({'tot_impuesto_pagar_x_percepcion': 'TOTAL IMPUESTO PAGAR POR PERCEPCION'}, inplace=True)
        # 1 df_pivot.replace({'total_pagado': 'TOTAL PAGADO'}, inplace=True)
        # 2 df_pivot.replace({'total_impuesto_a_pagar_2610': 'IMPUESTO A PAGAR'}, inplace=True)
        df_pivot.replace({'diferencia_arr_ct': 'DIFERENCIA ARRASTRE CREDITO TRIBUTARIO'}, inplace=True)
        df_pivot.replace({'diferencia_x_ct': 'DIFERENCIA POR CREDITO TRIBUTARIO'}, inplace=True)
        df_pivot.replace({'diferencia_adquisiciones': 'DIFERENCIA ADQUISICIONES'}, inplace=True)
        df_pivot.replace({'diferencia_retenciones': 'DIFERENCIA RETENCIONES'}, inplace=True)

        df_pivot.replace({'totales': 'TOTALES'}, inplace=True)
        df_pivot.replace({'codigo_impuesto': 'CODIGO IMPUESTO'}, inplace=True)
        df_pivot.replace({'calculo_ct_adq': 'CALCULO_CT_ADQ'}, inplace=True)
        df_pivot.replace({'calculo_ct_ret': 'CALCULO_CT_RET'}, inplace=True)
        self.dff = df_pivot.copy()

        rows = dataframe_to_rows(self.dff)

        worksheet_cea.sheet_properties.tabColor = '1072BA'
        worksheet_cea.merge_cells('E4:I4')
        worksheet_cea.merge_cells('E5:I5')
        worksheet_cea.merge_cells('F6:I6')
        worksheet_cea.merge_cells('F7:I7')
        worksheet_cea.merge_cells('F8:I8')
        if self.globales == 0:
            self.nav.nombre_analista = self.nav.nombre
            self.nav.nombre_supervisor = ''

        self.escribir_encabezado(worksheet_cea, '', self._sql.jd)

        ubicar_en_fila, ubicar_en_col = 12, 2
        for r_idx, row in enumerate(rows, 1):
            for c_idx, valor in enumerate(row, 1):
                if c_idx < 2:
                    continue
                if r_idx == 1:
                    self.escribir_celda(worksheet_cea, r_idx+ubicar_en_fila, c_idx - 1 + ubicar_en_col, valor, centrado=True, negrita=True, fondo=self.uf.sri, colorido=self.uf.blanco)
                else:
                    self.escribir_celda(worksheet_cea, r_idx - 1 + ubicar_en_fila, c_idx-1+ubicar_en_col, valor, centrado=False, negrita=False, fondo=self.uf.lavender, colorido=self.uf.negro)
        self.auto_ajustar(worksheet_cea)
        worksheet_cea.sheet_view.showGridLines = False
        worksheet_cea.sheet_view.zoomScale = 80
        if self.reporte_pie:
            self.escribir_footer(worksheet_cea, len(self.dff.index) + 5, 7, self.nav)
        return worksheet_cea

    def get_informe(self):
        '''informe'''
        self._sql.jd.procedencia = 'interna'
        workbook = Workbook()
        workbook.remove(workbook.active)
        sms = 'Reporte Analisis Previo Cadena'
        worksheet = workbook.create_sheet(title=sms, index=1)
        self.reporte_pie = True
        worksheet = self.get_hoja_analisis(worksheet)
        worksheet.title = f'Analisis Cadena ({len(self.dff.index)})'
        return workbook


class InformeDescartes_reporte(Pleyades.Abayo):
    ''' Reporte de Descartados:  '''
    def __init__(self, db):
        '''constructor principal'''
        super().__init__(db.uf)
        self.db = db
        self.uf = self.db.uf
        self.reporte_pie = False
        self.globales = 0
        self.df = botcito.DataFrame()
        self._sql = Reportes.Globales(self.uf.pi)
        self.nav = self.uf.navegante

    def get_hoja_cruces(self, worksheet):
        '''hoja de cruces'''
        self.db.uf.pi.ufuf = -17
        self._sql.jd.ufuf = -17
        info = Fotones.Reemplazantes(self.db)
        self.df = info.volcado(self._sql)

        df = self.df.reset_index()
        df.drop(columns=["index"], axis=1, inplace=True)
        df_pivot = df.reset_index()
        df_pivot.drop(columns=["index"], axis=1, inplace=True)

        self.df = df_pivot
        rows = dataframe_to_rows(self.df)
        worksheet.sheet_properties.tabColor = '1072BA'
        worksheet.merge_cells('E4:I4')
        worksheet.merge_cells('E5:I5')
        worksheet.merge_cells('F6:I6')
        worksheet.merge_cells('F7:I7')
        worksheet.merge_cells('F8:I8')
        if self.globales == 0:
            self.nav.nombre_analista = self.nav.nombre
            self.nav.nombre_supervisor = ''
        self.escribir_encabezado(worksheet, '', self._sql.jd)
        ubicar_en_fila, ubicar_en_col = 10, 4
        for r_idx, row in enumerate(rows, 1):
            for c_idx, valor in enumerate(row, 1):
                if (c_idx < 2):
                    continue
                if r_idx == 1:
                    self.escribir_celda(worksheet, r_idx+ubicar_en_fila, c_idx - 1 + ubicar_en_col, valor, centrado=True, negrita=True, fondo=self.uf.sri, colorido=self.uf.blanco)
                else:
                    self.escribir_celda(worksheet, r_idx - 1 + ubicar_en_fila, c_idx - 1 + ubicar_en_col, valor, centrado=False, negrita=False, fondo=self.uf.peach, colorido=self.uf.negro)

        self.auto_ajustar(worksheet)
        worksheet.sheet_view.showGridLines = False
        worksheet.sheet_view.zoomScale = 80
        if self.reporte_pie: 
            self.escribir_footer(worksheet, len(self.df.index), 6, self.nav)
        return worksheet

    def get_informe(self):
        '''informe intero'''
        self._sql.jd.procedencia = 'interna'
        workbook = Workbook()
        workbook.remove(workbook.active)
        worksheet = workbook.create_sheet(title='Filas Descartadas', index=1)
        self.reporte_pie = True
        worksheet = self.get_hoja_cruces(worksheet)
        worksheet.title = f'Descartes ({len(self.df.index)})'
        return workbook


class Responsiva():
    '''clase para retorno de resultados'''
    def __init__(self, abot, fecha, wb, pre):
        self.abot = abot
        self.fecha = fecha
        self.wb = wb
        self.pre = pre

    def devolver_response(self):
        '''devolver respuesta'''
        from io import BytesIO
        excel_stream = BytesIO()
        self.wb.save(excel_stream)
        excel_stream.seek(0)
        nombre = f"{self.abot}_{self.fecha}_{self.pre}.xlsx" 
        response = make_response(excel_stream.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename={nombre}'
        return response


class MultiLateral:
    '''multi origen'''
    def __init__(self, generador):
        self.db = generador.db
        self.cn = Consultas.Papel(self.db)
        self.jd = generador.db.uf.pi
        self._reporteador_ = generador
        diccc = {}
        _nav_ = Celebridades.Navegante(diccc)


class Continental:
    '''# 1 T = TypeVar("T", bound=Provireporte, covariant=True)'''
    '''# 2 class Continental(Generic[T]):'''

    def __init__(self, uni):
        '''constructor principal'''
        self.cn = Consultas.Papel(uni.db)
        self.rep = uni._reporteador_

    def get_fecha_ymd(self) -> str:
        '''fecha '''
        date_time = datetime.fromtimestamp(datetime.now().timestamp())
        return date_time.strftime("%Y-%m-%d %H:%M:%S")

    def get_sql_razon_social(self, contri):
        '''razon social'''
        if len(contri) == 10:
            contri = self.rep.uf.costelo(contri)
        rs = self.cn.get_sql_rs(contri)
        if len(rs.strip()) == 0:
            rs = self.cn.get_sql_rs(contri)
        return rs

    def crear_reporte(self):
        '''crear reporte'''
        abot = self.rep._sql.jd.contri
        if 10 == len(abot):
            abot = self.rep.uf.costelo((abot))
        if len(abot) == 13:
            self.rep._sql.jd.razon_social = \
                self.rep.db.get_scalar(self.get_sql_razon_social(abot))

        fecha = self.get_fecha_ymd()
        wb = self.rep.get_informe()
        res = Responsiva(abot, fecha, wb, self.rep._sql.jd.prefijo)
        return res.devolver_response()


class Puente():
    '''
        #20 de febrero 2024 - Se agrega reduccion con clase MultiLateral
        1 providencia
        2 cadena
        3 Informe de Revision , cruce de retenciones
        4 Reporte de Validaciones
        5 Reporte de resumen periodos
        6 Reporte Completo
    '''
    def __init__(self,  db):
        '''constructor principal'''
        self.db = db
        self.cn = Consultas.Papel(db)

    def reporte(self):
        '''reporte'''
        respuesta = ''
        if len(self.db.uf.pi.contri) == 13:
            self.db.uf.pi.razon_social = \
                self.db.get_scalar(self.cn.get_sql_razon_social())

        match (int(self.db.uf.pi.cuerda)):
            case 1:
                self.db.uf.pi.prefijo = 'PROV'
                hostel = Continental(MultiLateral(Provireporte(self.db)))
                respuesta = hostel.crear_reporte()

            case 2:
                self.db.uf.pi.prefijo = 'ANALISIS'
                hostel = Continental(MultiLateral(Cadenareporte(self.db)))
                respuesta = hostel.crear_reporte()

            case 3:
                self.db.uf.pi.prefijo = 'INFR'
                hostel = Continental(MultiLateral(InformeRevreporte(self.db)))
                respuesta = hostel.crear_reporte()

            case 4:
                self.db.uf.pi.prefijo = 'VALD'
                hostel = \
                    Continental(MultiLateral(Validacionesreporte(self.db)))
                respuesta = hostel.crear_reporte()

            case 5:
                self.db.uf.pi.prefijo = 'RESPE'
                hostel = Continental(MultiLateral(Resumen_pers_repo(self.db)))
                respuesta = hostel.crear_reporte()

            case 12:
                self.db.uf.pi.prefijo = 'DECLA'
                hostel = \
                    Continental(MultiLateral(InformeDeclas_reporte(self.db)))
                respuesta = hostel.crear_reporte()
            case 15:
                self.db.uf.pi.prefijo = 'FUTU'
                hostel = \
                    Continental(MultiLateral(InformeFuturas_reporte(self.db)))
                respuesta = hostel.crear_reporte()

            case 16:
                self.db.uf.pi.prefijo = 'ANALISIS PRE'
                hostel = \
                    Continental(MultiLateral(InformeCadenaEarlier_reporte(self.db)))
                respuesta = hostel.crear_reporte()

            case 17:
                self.db.uf.pi.prefijo = 'DESCARTADOS'
                hostel = \
                    Continental(MultiLateral(InformeDescartes_reporte(self.db)))
                respuesta = hostel.crear_reporte()

            case 6 | 7 | 8 | 9:
                print(f" opcion  {self.db.uf.pi.cuerda} periodo ini = {self.db.uf.pi.periodo_inicial}  periodo fin = {self.db.uf.pi.periodo_final}")
                self.db.uf.pi.prefijo = 'GLOBAL'
                hostel = Continental(MultiLateral(Globalreporte(self.db)))
                respuesta = hostel.crear_reporte()

        return respuesta

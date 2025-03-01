"""Consultas, desde Enero 2023
Funcionalidades:
  - Consultas a la base para postgres.
ESTANDAR PEP8
"""


from openpyxl.styles import PatternFill
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Protection
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime


class Abayo:
    '''para escribir en la hoja de excel'''
    def __init__(self, uf):
        '''constructor principal'''
        self.uf = uf

    def escribir_celda(self,
                       worksheet,
                       row_num,
                       col_num, valor, centrado, negrita, fondo, colorido):
        '''escritura de la celda'''
        cell = worksheet.cell(row=row_num, column=col_num, value=valor)
        if centrado:
            cell.alignment = Alignment(horizontal='center',
                                       vertical='center',
                                       wrap_text=True)
        else:
            cell.alignment = Alignment(horizontal='left',
                                       vertical='center',
                                       wrap_text=True)
        if negrita:
            cell.font = Font(name="Arial Narrow",
                             size=8,
                             bold=True,
                             color=colorido)  
        else:
            cell.font = Font(name="Arial Narrow", size=8, bold=False)
            cell.number_format = '#,##0.00'
        if fondo != '':
            cell.fill = PatternFill(start_color=fondo,
                                    end_color=fondo,
                                    fill_type="solid")

    def auto_ajustar(self, worksheet):
        '''auto ajuste de las celdas'''
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            worksheet.column_dimensions[column_letter].width = adjusted_width

    def escribir_encabezado(self, worksheet,  titulo, _jd):
        '''escritura del encabezado'''
        worksheet.merge_cells('E4:I4')
        worksheet.merge_cells('E5:I5')
        worksheet.merge_cells('F6:I6')
        worksheet.merge_cells('F7:I7')
        worksheet.merge_cells('F8:I8')
        if _jd.contri == 10:
            self.uf.costelo(_jd.contri)
        self.escribir_celda(worksheet, 4, 5, "SERVICIO DE RENTAS INTERNAS", centrado=True, negrita=True, fondo= self.uf.sri, colorido=self.uf.blanco)
        self.escribir_celda(worksheet, 5, 5, titulo, centrado=True, negrita=True, fondo=self.uf.sri, colorido=self.uf.blanco)
        self.escribir_celda(worksheet, 6, 5, "No. RUC:", centrado=False, negrita=True, fondo=self.uf.blanco, colorido=self.uf.negro)
        self.escribir_celda(worksheet, 7, 5, "RAZÓN SOCIAL:", centrado=False, negrita=True, fondo=self.uf.blanco, colorido=self.uf.negro)
        self.escribir_celda(worksheet, 8, 5, "PERÍODO SOLICITADO:", centrado=False, negrita=True, fondo=self.uf.blanco, colorido=self.uf.negro)
        self.escribir_celda(worksheet, 9, 5, "TRÁMITE No.", centrado=False, negrita=True, fondo=self.uf.blanco, colorido=self.uf.negro)
        self.escribir_celda(worksheet, 6, 6, _jd.contri, centrado=False, negrita=True, fondo=self.uf.blanco, colorido=self.uf.negro)
        self.escribir_celda(worksheet, 7, 6, _jd.razon_social, centrado=False, negrita=True, fondo=self.uf.blanco, colorido=self.uf.negro)
        self.escribir_celda(worksheet, 8, 6, _jd.periodo_inicial[0:7] + '  -  ' + _jd.periodo_final[0:7], centrado=False, negrita=True, fondo=self.uf.blanco, colorido=self.uf.negro)
        self.escribir_celda(worksheet, 9, 6, _jd.tramite, centrado=False, negrita=True, fondo=self.uf.blanco, colorido=self.uf.negro)

    def get_fecha_ymd_hms(self) -> str:
        ''' se obtiene la fecha hora mes y dia'''
        date_time = datetime.fromtimestamp(datetime.now().timestamp())
        return date_time.strftime("%Y-%m-%d %H:%M:%S")

    def escribir_footer(self, ws_periodos, longitud_dataframe,
                        empieza_en_col, nav):
        '''#pie de pagina '''
        r_idx = 0
        c_idx = 0
        ubicar_en_fila, ubicar_en_col = 20 + longitud_dataframe, empieza_en_col
        self.escribir_celda(ws_periodos, r_idx + ubicar_en_fila, c_idx + ubicar_en_col, "Elaborado Por", centrado=True, negrita=True, fondo=self.uf.sri, colorido=self.uf.blanco)
        c_idx += 1
        self.escribir_celda(ws_periodos, r_idx + ubicar_en_fila, c_idx + ubicar_en_col, nav.nombre_analista , centrado=True, negrita=True, fondo=self.uf.blanco, colorido=self.uf.negro)
        r_idx += 1
        c_idx = 0
        self.escribir_celda(ws_periodos, r_idx + ubicar_en_fila, c_idx + ubicar_en_col, "Firma", centrado=True, negrita=True, fondo=self.uf.sri, colorido=self.uf.blanco)

        ubicar_en_fila, ubicar_en_col = 24 + longitud_dataframe, empieza_en_col 
        self.escribir_celda(ws_periodos, r_idx + ubicar_en_fila, c_idx + ubicar_en_col, "Revisado Por", centrado=True, negrita=True, fondo=self.uf.sri, colorido=self.uf.blanco)
        c_idx += 1
        self.escribir_celda(ws_periodos, r_idx + ubicar_en_fila, c_idx + ubicar_en_col, nav.nombre_supervisor , centrado=True, negrita=True, fondo=self.uf.blanco, colorido=self.uf.negro)
        c_idx = 0
        r_idx += 1
        self.escribir_celda(ws_periodos, r_idx + ubicar_en_fila, c_idx + ubicar_en_col, "Firma", centrado=True, negrita=True, fondo=self.uf.sri, colorido=self.uf.blanco)
        r_idx += 2
        self.escribir_celda(ws_periodos, r_idx + ubicar_en_fila, c_idx + ubicar_en_col, "Fecha", centrado=True, negrita=True, fondo=self.uf.sri, colorido=self.uf.blanco)
        c_idx += 1
        self.escribir_celda(ws_periodos, r_idx + ubicar_en_fila, c_idx + ubicar_en_col, self.get_fecha_ymd_hms() , centrado=True, negrita=True, fondo=self.uf.blanco, colorido=self.uf.negro)

    def escribir_en_red(self, fila, worksheet, r_idx, c_idx, campo,
                        campo_sombra):
        '''escritura en rojo'''
        blanco = '00FFFFFF'
        rojo = '00FF0000'
        negro = '00000000'
        fondo = blanco
        if fila[campo_sombra] == 0:
            fondo = rojo
        self.escribir_celda(worksheet,
                            r_idx,
                            c_idx,
                            fila[campo],
                            centrado=True,
                            negrita=True, fondo=fondo, colorido=negro)
